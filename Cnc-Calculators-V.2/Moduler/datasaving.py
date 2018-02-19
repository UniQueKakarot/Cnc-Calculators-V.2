import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os


class CuttingSpeedData():

    def __init__(self, data):

        """ Checking if the xlsx file has all the necessities,
            and if not generate it """

        self.data = data

        self.checkfile(self.data)

        self.wb = load_workbook(self.data)

        test = 0

        for i in self.wb.get_sheet_names():
            if i == "Cutting Speed":
                test = 1

        if test == 0:
            self.ws = self.wb.create_sheet("Cutting Speed")

            try:
                x = self.wb.get_sheet_by_name('Sheet')
                self.wb.remove_sheet(x)
            except KeyError:
                print("Worksheet does not exist, carrying on.")

            ws = self.wb["Cutting Speed"]
            ws.cell(row=2, column=2, value="Cutting Meter")
            ws.cell(row=2, column=3, value="Mill Diameter")
            ws.cell(row=2, column=4, value="Number of teeth")
            ws.cell(row=2, column=5, value="Feed pr Tooth")

            self.wb.save(self.data)

    def filesave(self, cuttingdata, mill, teeth, tooth):

        """ Collecting what is written in the textinput boxes for the
            cuttingspeed module and saving it to a xlsx file """

        ws = self.wb["Cutting Speed"]

        row = 3
        while ws.cell(row=row, column=2).value != None:
            row += 1

        try:
            cuttingdata = int(cuttingdata)
        except ValueError:
            cuttingdata = 0
        try:
            mill = mill.replace(',', '.')
            mill = float(mill)
        except ValueError:
            mill = 0
        try:
            teeth = int(teeth)
        except ValueError:
            teeth = 0
        try:
            tooth = tooth.replace(',', '.')
            tooth = float(tooth)
        except ValueError:
            tooth = 0

        ws.cell(row=row, column=2, value=cuttingdata)
        ws.cell(row=row, column=3, value=mill)
        ws.cell(row=row, column=4, value=teeth)
        ws.cell(row=row, column=5, value=tooth)

        try:
            self.wb.save(self.data)
        except PermissionError:
            print("Can't access file, please close if open!")

    def checkfile(self, data):

        """ Method for checking if the file exists and if not, generate new """

        if os.path.isfile(data) is True:
            pass
        else:
            wb = Workbook()
            wb.save(data)


class MaterialRemovalData():

    def __init__(self, data):

        """ Checking if the xlsx file has all the necessities,
            and if not generate it """

        self.data = data

        self.checkfile(self.data)

        self.wb = load_workbook(self.data)

        test = 0

        for i in self.wb.get_sheet_names():
            if i == "Material Removal Rate":
                test = 1

        if test == 0:
            self.ws = self.wb.create_sheet("Material Removal Rate")

            try:
                x = self.wb.get_sheet_by_name('Sheet')
                self.wb.remove_sheet(x)
            except KeyError:
                print("Worksheet does not exist, carrying on.")

            ws = self.wb["Material Removal Rate"]
            ws.cell(row=2, column=2, value="Depth of cut")
            ws.cell(row=2, column=3, value="Width of cut")
            ws.cell(row=2, column=4, value="Feedrate")
            ws.cell(row=2, column=5, value="Material Removal Rate")

            self.wb.save(self.data)

    def filesave(self, depthcut, widthcut, feedrate, mrr):

        """ Collecting what is written in the textinput boxes for the
            cuttingspeed calc and saving it to a xlsx file """

        ws = self.wb["Material Removal Rate"]

        row = 3
        while ws.cell(row=row, column=2).value != None:
            row += 1

        try:
            depthcut = depthcut.replace(',', '.')
            depthcut = float(depthcut)
        except ValueError:
            depthcut = 0
        try:
            widthcut = widthcut.replace(',', '.')
            widthcut = float(widthcut)
        except ValueError:
            widthcut = 0
        try:
            feedrate = int(feedrate)
        except ValueError:
            feedrate = 0

        ws.cell(row=row, column=2, value=depthcut)
        ws.cell(row=row, column=3, value=widthcut)
        ws.cell(row=row, column=4, value=feedrate)

        if mrr == "Please input values":
            ws.cell(row=row, column=5, value=0)
        else:
            ws.cell(row=row, column=5, value=mrr)

        ws.cell(row=row, column=6, value='cm³/min')

        try:
            self.wb.save(self.data)
        except PermissionError:
            print("Can't access file, please close if open!")

    def checkfile(self, data):

        """ Method for checking if the file exists and if not, generate new """

        if os.path.isfile(data) is True:
            pass
        else:
            wb = Workbook()
            wb.save(data)


class HelixData():

    def __init__(self, data):

        """ Checking if the xlsx file has all the necessities,
            and if not generate it """

        self.data = data

        self.checkfile(self.data)

        self.wb = load_workbook(self.data)

        test = 0

        for i in self.wb.get_sheet_names():
            if i == "Helix Angle":
                test = 1

        if test == 0:
            self.ws = self.wb.create_sheet("Helix Angle")

            try:
                x = self.wb.get_sheet_by_name('Sheet')
                self.wb.remove_sheet(x)
            except KeyError:
                print("Worksheet does not exist, carrying on.")

            ws = self.wb["Helix Angle"]
            ws.cell(row=2, column=2, value="Mill Diameter")
            ws.cell(row=2, column=3, value="Hole Diameter")
            ws.cell(row=2, column=4, value="Step/Pitch")
            ws.cell(row=2, column=5, value="Angle of Decent")

            self.wb.save(self.data)

    def filesave(self, milldia, holedia, zstep, angle):

        """ Collecting what is written in the textinput boxes for the
            Spiral module and saving it to a xlsx file """

        ws = self.wb["Helix Angle"]

        row = 3
        while ws.cell(row=row, column=2).value != None:
            row += 1

        try:
            milldia = milldia.replace(',', '.')
            milldia = float(milldia)
        except ValueError:
            milldia = 0
        try:
            holedia = holedia.replace(',', '.')
            holedia = float(holedia)
        except ValueError:
            holedia = 0
        try:
            zstep = zstep.replace(',', '.')
            zstep = float(zstep)
        except ValueError:
            zstep = 0

        ws.cell(row=row, column=2, value=milldia)
        ws.cell(row=row, column=3, value=holedia)
        ws.cell(row=row, column=4, value=zstep)

        if angle == "Please input values":
            ws.cell(row=row, column=5, value=0)
        else:
            ws.cell(row=row, column=5, value=angle)

        ws.cell(row=row, column=6, value='°')

        try:
            self.wb.save(self.data)
        except PermissionError:
            print("Can't access file, please close if open!")

    def checkfile(self, data):

        """ Method for checking if the file exists and if not, generate new """

        if os.path.isfile(data) is True:
            pass
        else:
            wb = Workbook()
            wb.save(data)


class RampData():

    def __init__(self, data):

        """ Checking if the xlsx file has all the necessities,
            and if not generate it """

        self.data = data

        self.checkfile(self.data)

        self.wb = load_workbook(self.data)

        test = 0

        for i in self.wb.get_sheet_names():
            if i == "Ramp Angle":
                test = 1

        if test == 0:
            self.ws = self.wb.create_sheet("Ramp Angle")

            try:
                x = self.wb.get_sheet_by_name('Sheet')
                self.wb.remove_sheet(x)
            except KeyError:
                print("Worksheet does not exist, carrying on.")

            ws = self.wb["Ramp Angle"]
            ws.cell(row=2, column=2, value="Toolpath Length")
            ws.cell(row=2, column=3, value="Step/Pitch")
            ws.cell(row=2, column=4, value="Angle of Decent")

            self.wb.save(self.data)

    def filesave(self, tpl, zstep, angle):

        """ Collecting what is written in the textinput boxes for the
            Ramp Angle module and saving it to a xlsx file """

        ws = self.wb["Ramp Angle"]

        row = 3
        while ws.cell(row=row, column=2).value != None:
            row += 1

        try:
            tpl = tpl.replace(',', '.')
            tpl = float(tpl)
        except ValueError:
            tpl = 0
        try:
            zstep = zstep.replace(',', '.')
            zstep = float(zstep)
        except ValueError:
            zstep = 0

        ws.cell(row=row, column=2, value=tpl)
        ws.cell(row=row, column=3, value=zstep)

        if angle == "Please input values":
            ws.cell(row=row, column=4, value=0)
        else:
            ws.cell(row=row, column=4, value=angle)

        ws.cell(row=row, column=5, value='°')

        try:
            self.wb.save(self.data)
        except PermissionError:
            print("Can't access file, please close if open!")

    def checkfile(self, data):

        """ Method for checking if the file exists and if not, generate new """

        if os.path.isfile(data) is True:
            pass
        else:
            wb = Workbook()
            wb.save(data)


class SurfaceRaData():

    def __init__(self, data):

        """ Checking if the xlsx file has all the necessities,
            and if not generate it """

        self.data = data

        self.checkfile(self.data)

        self.wb = load_workbook(self.data)

        test = 0

        for i in self.wb.get_sheet_names():
            if i == "Surface Roughness":
                test = 1

        if test == 0:
            self.ws = self.wb.create_sheet("Surface Roughness")

            try:
                x = self.wb.get_sheet_by_name('Sheet')
                self.wb.remove_sheet(x)
            except KeyError:
                print("Worksheet does not exist, carrying on.")

            ws = self.wb["Surface Roughness"]
            ws.cell(row=2, column=2, value="Feed per Turn")
            ws.cell(row=2, column=3, value="Nose Radius")
            ws.cell(row=2, column=4, value="Ra")

            self.wb.save(self.data)

    def filesave(self, feed, nr, ra):

        """ Collecting what is written in the textinput boxes for the
            Ra module and saving it to a xlsx file """

        ws = self.wb["Surface Roughness"]

        row = 3
        while ws.cell(row=row, column=2).value != None:
            row += 1

        try:
            feed = feed.replace(',', '.')
            feed = float(feed)
        except ValueError:
            feed = 0
        try:
            nr = nr.replace(',', '.')
            nr = float(nr)
        except ValueError:
            nr = 0

        ws.cell(row=row, column=2, value=feed)
        ws.cell(row=row, column=3, value=nr)

        if ra == "Please input values":
            ws.cell(row=row, column=4, value=0)
        else:
            ws.cell(row=row, column=4, value=ra)

        try:
            self.wb.save(self.data)
        except PermissionError:
            print("Can't access file, please close if open!")

    def checkfile(self, data):

        """ Method for checking if the file exists and if not, generate new """

        if os.path.isfile(data) is True:
            pass
        else:
            wb = Workbook()
            wb.save(data)
